#!/usr/bin/env python3
"""
xlsx_to_tsv.py
==============
大きな発現行列の .xlsx / .xls を TSV に変換する。進捗を行数で表示する。

なぜ必要か:
  60,660遺伝子 × 1,211検体 = 約7,300万セルの .xlsx を pandas.read_excel で
  読むと、openpyxlがXMLを1セルずつ解析するため20〜60分かかり、しかも
  進捗が出ないので固まったように見える。同じファイルを読み直すたびに
  この時間を払うことになる。

  一度TSVにしておけば、以降の読み込みは数十秒で済む。
  preprocess_tcga.py は .tsv をそのまま受け付けるので、変換後は
  --new-tpm などにTSVのパスを渡すだけでよい。

使い方:

    python3 xlsx_to_tsv.py TCGA_BRCA_TPM.xlsx
        -> TCGA_BRCA_TPM.tsv を同じ場所に作る

    python3 xlsx_to_tsv.py input.xlsx --output /path/to/out.tsv

    # まとめて変換
    for f in *.xlsx *.xls; do python3 xlsx_to_tsv.py "$f"; done

.xlsx は1行ずつ流し読みするのでメモリをほとんど使わない。
.xls(旧形式)は仕様上ストリーミングできないため、pandas経由で一括読み込みする
(旧形式は65,536行までなので、実用上は問題にならない)。
"""

import argparse
import sys
import time
from pathlib import Path


def convert_xlsx(src: Path, dst: Path, progress_every: int) -> int:
    """openpyxl の read_only モードで1行ずつ書き出す(メモリ非依存)。"""
    from openpyxl import load_workbook

    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    total = ws.max_row or 0
    print(f"  シート '{ws.title}' / 推定 {total:,} 行 × {ws.max_column or 0:,} 列")

    started = time.time()
    n = 0
    with dst.open("w", encoding="utf-8", newline="") as out:
        for row in ws.iter_rows(values_only=True):
            # 末尾の空セルを落とす(xlsxは行ごとに列数が揃わないことがある)
            cells = list(row)
            while cells and cells[-1] is None:
                cells.pop()
            out.write("\t".join("" if c is None else str(c) for c in cells))
            out.write("\n")
            n += 1
            if n % progress_every == 0:
                elapsed = time.time() - started
                rate = n / elapsed if elapsed else 0
                eta = (total - n) / rate if rate and total > n else 0
                print(f"    {n:,} 行 ({elapsed:,.0f}秒経過, "
                      f"{rate:,.0f} 行/秒, 残り約 {eta / 60:,.1f} 分)", flush=True)
    wb.close()
    return n


def convert_xls(src: Path, dst: Path) -> int:
    """旧形式(.xls)。ストリーミングできないので一括で読む。"""
    import pandas as pd
    print("  .xls は一括読み込みします(進捗は出ません)…", flush=True)
    df = pd.read_excel(src)
    df.to_csv(dst, sep="\t", index=False)
    return len(df) + 1


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", help="変換元の .xlsx / .xls")
    ap.add_argument("--output", help="出力先(省略時は拡張子を .tsv に変えたもの)")
    ap.add_argument("--progress-every", type=int, default=2000,
                    help="何行ごとに進捗を出すか(既定: 2000)")
    ap.add_argument("--overwrite", action="store_true",
                    help="出力先が既にある場合に上書きする")
    args = ap.parse_args()

    src = Path(args.input)
    if not src.is_file():
        sys.exit(f"ファイルがありません: {src}")
    dst = Path(args.output) if args.output else src.with_suffix(".tsv")
    if dst.exists() and not args.overwrite:
        sys.exit(f"{dst} は既に存在します。上書きするなら --overwrite を付けてください。")

    print(f"{src.name} ({src.stat().st_size / 1e6:,.0f} MB) -> {dst}")
    started = time.time()

    if src.suffix.lower() == ".xlsx":
        n = convert_xlsx(src, dst, args.progress_every)
    elif src.suffix.lower() == ".xls":
        n = convert_xls(src, dst)
    else:
        sys.exit(f"対応していない拡張子です: {src.suffix}")

    elapsed = time.time() - started
    print(f"\n完了: {n:,} 行 / {elapsed / 60:,.1f} 分 -> {dst} "
          f"({dst.stat().st_size / 1e6:,.0f} MB)")
    print("preprocess_tcga.py にはこのTSVのパスを渡してください。")


if __name__ == "__main__":
    main()
